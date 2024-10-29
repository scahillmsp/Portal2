import openpyxl
from django.core.management.base import BaseCommand
from shop.models import Product, Vehicle
from django.db import IntegrityError, transaction

class Command(BaseCommand):
    help = 'Imports products from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file.')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']

        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Get the total number of rows to calculate the percentage
        total_rows = sheet.max_row - 1  # Subtract 1 to account for the header row

        # Start a database transaction to ensure consistency
        with transaction.atomic():
            # Loop through rows in the sheet and import products
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):  # Skip the header row
                try:
                    part_number, oem_number, header, description, price, stock, image, category, make, model, version = row

                    # Ensure a default price if missing
                    if price is None:
                        price = 0.00  # Default price to 0.00 if missing

                    # Ensure stock is set, default to 0 if missing
                    if stock is None:
                        stock = 0

                    # Ensure version is treated as a string, default to "Unknown Version" if missing
                    if version is None or str(version).strip() == '':
                        version = "Unknown Version"  # Default version

                    # Create or get the Vehicle instance
                    vehicle, created = Vehicle.objects.get_or_create(
                        category=category,
                        make=make,
                        model=model,
                        version=version
                    )

                    # Use update_or_create to either update or create the Product
                    product, created = Product.objects.update_or_create(
                        part_number=part_number,
                        defaults={
                            'oem_number': oem_number,
                            'header': header,
                            'description': description,
                            'price': price,
                            'stock': stock,
                            'image': image
                        }
                    )

                    # Add the vehicle compatibility to the product
                    product.vehicles.add(vehicle)

                    # Calculate and display the progress
                    percentage = (index / total_rows) * 100
                    self.stdout.write(self.style.SUCCESS(f'Processed {index} of {total_rows} rows ({percentage:.2f}% complete)'))

                except IntegrityError as e:
                    self.stdout.write(self.style.ERROR(f"Error on row {index}: {e}"))

        self.stdout.write(self.style.SUCCESS('Successfully imported all products'))
